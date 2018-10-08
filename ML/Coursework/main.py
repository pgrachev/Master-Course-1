from openpyxl import load_workbook
from keras.models import Sequential
from keras.layers import Dense
from sklearn.metrics import accuracy_score
from sklearn.svm import SVC
from sklearn.naive_bayes import GaussianNB
from sklearn import model_selection
from sklearn.neighbors import KNeighborsClassifier
from sklearn.tree import DecisionTreeClassifier
from sklearn.discriminant_analysis import LinearDiscriminantAnalysis
import matplotlib
from pandas.plotting import scatter_matrix
import matplotlib.pyplot as plt
from sklearn.ensemble import ExtraTreesClassifier
import sklearn.linear_model as lm
import pandas as pd
import numpy as np
import os
from keras.callbacks import History

N_FEATURES = 13
N_CLASSES = 16
MAIN_FEATURES = 4
EXAMPLES = 50
FIRST_ROW = 3
FIRST_COL = 9

N_HIDDEN = 16
BATCH_SIZE = 32
N_EPOCHS = 1000

accuracies = []
wb = load_workbook('./SyringaDate_2017.xlsx')

sheet = wb['Лист1']

X, Y, vY = [], [], []
FEATURE_NAMES = []


def int_to_vec(lst):
    result = []
    for i in lst:
        res = np.zeros(N_CLASSES)
        res[i] = 1
        result.append(res)
    return np.array(result)


for cl in range(N_CLASSES):
    fst_row = cl * (N_FEATURES + 1) + FIRST_ROW
    lst_row = fst_row + N_FEATURES
    for col in range(FIRST_COL, FIRST_COL + EXAMPLES):
        TX = []
        for row in range(fst_row, lst_row):
            TX.append(sheet.cell(row = row, column = col).value)
        X.append(np.array(TX))
        Y.append(cl)

X = np.array(X)
Y = np.array(Y)




for i in range(N_FEATURES):
    FEATURE_NAMES.append(sheet.cell(row = FIRST_ROW + i, column = 6).value)
print(FEATURE_NAMES)

df = pd.DataFrame.from_records(X, columns=FEATURE_NAMES)

forest = ExtraTreesClassifier(n_estimators=250, random_state=239)
forest.fit(X, Y)
importances = forest.feature_importances_
std = np.std([tree.feature_importances_ for tree in forest.estimators_],
             axis=0)
indices = np.argsort(importances)[::-1]

f = open('feature_importance_ranking.txt', 'w', encoding='utf-8')
important_features = []

for i in range(X.shape[1]):
    f.write(str(i + 1) + '. ' + FEATURE_NAMES[indices[i]] + ': ' + str(round(importances[indices[i]], 3)) + '\n')
    if (i < MAIN_FEATURES):
        important_features.append(FEATURE_NAMES[indices[i]])
f.close()

df_important = df[important_features]
matplotlib.rcParams.update({'font.size': 8})
sm = scatter_matrix(df_important)
[s.yaxis.label.set_rotation(80) for s in sm.reshape(-1)]
plt.savefig('Scatterplot Matrix.png')
plt.close()

if not os.path.exists('boxes'):
    os.makedirs('boxes')

for i in range(N_FEATURES):
    df[FEATURE_NAMES[i]].plot(kind='box', subplots=True, layout=(1, 1), sharex=False, sharey=False)
    plt.savefig('boxes/' + FEATURE_NAMES[i] + '.png')
    plt.close()

if not os.path.exists('hists'):
    os.makedirs('hists')

for i in range(N_FEATURES):
    df[FEATURE_NAMES[i]].hist()
    plt.savefig('hists/' + FEATURE_NAMES[i] + '.png')
    plt.close()

models = []
models.append(('Logistic Regression', lm.LogisticRegression()))
models.append(('Linear Discriminant Analysis', LinearDiscriminantAnalysis()))
models.append(('K-Neighbors Classifier', KNeighborsClassifier()))
models.append(('Decision Tree Classifier', DecisionTreeClassifier()))
models.append(('Gaussian Naive Bayes', GaussianNB()))
models.append(('Support Vector Machine', SVC()))

results = []
names = []
for name, model in models:
    kfold = model_selection.KFold(n_splits=8, shuffle=True, random_state=239)
    cv_results = model_selection.cross_val_score(model, X, Y, cv=kfold, scoring='accuracy')
    results.append(cv_results)
    names.append(name)
    accuracies.append([name, cv_results.mean(), cv_results.std()])
    msg = "%s: %f (%f)" % (name, cv_results.mean(), cv_results.std())
    print(msg)


losses = np.zeros(N_EPOCHS)
results = []
kfold = model_selection.KFold(n_splits=8, shuffle=True, random_state=239)
for tri, tei in kfold.split(X):
    X_train, X_test = X[tri], X[tei]
    y_train, y_test = int_to_vec(Y[tri]), Y[tei]
    history = History()
    model = Sequential()
    model.add(Dense(N_HIDDEN, activation='sigmoid', input_dim=13))
    model.add(Dense(N_CLASSES, activation='softmax'))
    model.compile(optimizer='adam',
                  loss='categorical_crossentropy',
                  metrics=['accuracy'])
    history = model.fit(X_train, y_train, epochs=N_EPOCHS, batch_size=BATCH_SIZE)
    losses = losses + np.array(history.history['loss'])
    predictions = model.predict(X_test)
    inds = np.array([np.argmax(pred) for pred in predictions])
    results.append(accuracy_score(y_test, inds))
losses = losses / kfold.get_n_splits()
xs = np.arange(N_EPOCHS)
results = np.array(results)
plt.plot(xs, losses, 'r')
plt.xlabel('Epochs')
plt.ylabel('Loss')
plt.savefig('loss_func_plot.png')
msg = "%s: %f (%f)" % ('Neural Network', results.mean(), results.std())
print(msg)
accuracies.append(['Neural Network', results.mean(), results.std()])
f = open('accuracies_comparision.txt', 'w')
f.write('Method / Mean Accuracy / Standart deviation\n')
for meth in accuracies:
    f.write(meth[0] + ' / ' + str(round(meth[1], 4)) + ' / ' + str(round(meth[2], 4)) + '\n')
